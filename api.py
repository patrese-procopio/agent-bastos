import uvicorn
import os
import glob
import json
import io
import wave
import tempfile
import re
import threading
from datetime import datetime, timedelta
from fastapi import Form, FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel, field_validator
from groq import Groq
from modules.rag import conversar_com_bastos, conversar_com_fontes
from modules.decifrar import transcrever_documento_bytes, TipoDocumento
from api_liderancas_router import liderancas_router
from routers.alertas_router import router as alertas_router
from routers.dashboard_router import router as dashboard_router
from routers.transcricao_router import router as transcricao_router
from services.alertas_service import (
    ler_alertas    as _ler_alertas,
    salvar_alertas as _salvar_alertas,
    ALERTAS_PATH       as _ALERTAS_PATH,
    ALERTAS_OSINT_PATH as _ALERTAS_OSINT_PATH,
)
from modules.liderancas import (
    ESTRUTURA, FACCOES, CARGOS_POR_FACCAO, estrutura_com_celas,
    criar_lider, atualizar_lider, deletar_lider,
    buscar_lider, listar_por_unidade,
    salvar_foto, carregar_foto,
)


# â”€â”€â”€ ConfiguraÃ§Ã£o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
PASTA_RELATORIOS = os.path.join(BASE_DIR, "data", "relatorios")
_SA_KEY_PATH     = os.path.join(BASE_DIR, "serviceAccountKey.json")

# Cria diretÃ³rios necessÃ¡rios uma Ãºnica vez na inicializaÃ§Ã£o
os.makedirs(PASTA_RELATORIOS, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "data", "audios"), exist_ok=True)

# Singleton Groq â€” uma Ãºnica conexÃ£o reutilizada em todas as requests
_GROQ_API_KEY = os.getenv("GROQ_API_KEY")
if not _GROQ_API_KEY:
    raise RuntimeError("GROQ_API_KEY nÃ£o encontrada. Configure o arquivo .env.")
_groq = Groq(api_key=_GROQ_API_KEY)

# Limites de seguranÃ§a
_MAX_PERGUNTA    = 4_000
_MAX_AUDIO_BYTES = 25 * 1024 * 1024
_AUDIO_EXTS = {".wav", ".mp3", ".mp4", ".ogg", ".webm", ".flac", ".m4a", ".mpga", ".mpeg"}

# Lock para escrita de arquivos de alertas (requests simultÃ¢neas)
_alertas_lock = threading.Lock()

_MESES_PT = [
    "Janeiro", "Fevereiro", "MarÃ§o", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]
_EXPORT_MIME = {
    "txt":  "text/plain",
    "pdf":  "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

_ALERTAS_PATH         = os.path.join(PASTA_RELATORIOS, "alertas.json")
_ALERTAS_OSINT_PATH   = os.path.join(PASTA_RELATORIOS, "alertas_osint.json")
_DASHBOARD_STATS_PATH = os.path.join(PASTA_RELATORIOS, "producao.json")
_INDICE_PATH          = os.path.join(BASE_DIR, "indice_documentos.json")

# â”€â”€â”€ Lista Negra â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_LISTA_NEGRA_FILE_ID = "1G6eFhb0jnD38iWU_SLDIFkGtOB0ngHjh"
_GDRIVE_SCOPES       = ["https://www.googleapis.com/auth/drive.readonly"]


def _seed_alertas_iniciais() -> None:
    now = datetime.now()

    if not os.path.exists(_ALERTAS_PATH):
        alertas = [
            {
                "id": "a1", "tipo": "telegram", "fonte": "@manausnoticias",
                "link": "https://t.me/manausnoticias",
                "titulo": "MenÃ§Ã£o via #CVAM",
                "resumo": "MovimentaÃ§Ã£o no bairro Compensa. Fonte relata presenÃ§a de elemento "
                          "conhecido como 'CarnaÃºba' coordenando distribuiÃ§Ã£o de entorpecentes. #CVAM #Manaus",
                "risco": "ALTO",
                "timestamp": (now - timedelta(minutes=18)).isoformat(),
                "lido": False,
                "alvo_id": 1, "alvo_nome": "Gelson de Lima CarnaÃºba", "alvo_vulgos": ["CarnaÃºba"],
                "termo_encontrado": "carnaÃºba", "hashtag": "#CVAM",
                "analise_ia": "Elemento identificado em Ã¡rea de distribuiÃ§Ã£o ativa â€” risco operacional "
                              "imediato. VerificaÃ§Ã£o de campo recomendada nas prÃ³ximas 2h.",
            },
            {
                "id": "b2", "tipo": "noticia", "fonte": "G1 Amazonas",
                "link": "https://g1.globo.com/am/",
                "titulo": "PolÃ­cia prende suspeito de trÃ¡fico no Jorge Teixeira",
                "resumo": '"John Wick" foi detido com 3 kg de entorpecentes durante operaÃ§Ã£o da DENARC nesta manhÃ£.',
                "risco": "ALTO",
                "timestamp": (now - timedelta(minutes=45)).isoformat(),
                "lido": False,
                "alvo_id": 14, "alvo_nome": "Leandro Costa de Oliveira",
                "alvo_vulgos": ["John Wick", "GaviÃ£o", "Leandrinho"],
                "termo_encontrado": "john wick",
                "analise_ia": "ConfirmaÃ§Ã£o de prisÃ£o â€” atualizar status do alvo. Verificar mandados pendentes.",
            },
            {
                "id": "c3", "tipo": "telegram", "fonte": "@policiaamazonas",
                "link": "https://t.me/policiaamazonas",
                "titulo": 'MenÃ§Ã£o: "El Diablo" em @policiaamazonas',
                "resumo": 'Alerta de fronteira: indivÃ­duo colombiano "El Diablo" teria cruzado pelo municÃ­pio de Tabatinga.',
                "risco": "ALTO",
                "timestamp": (now - timedelta(hours=1, minutes=30)).isoformat(),
                "lido": True,
                "alvo_id": 30, "alvo_nome": "Nelson Gaviria Florez",
                "alvo_vulgos": ["El Diablo", "Diablo"],
                "termo_encontrado": "el diablo",
                "analise_ia": None,
            },
            {
                "id": "d4", "tipo": "noticia", "fonte": "AcrÃ­tica AM",
                "link": "https://www.acritica.com/",
                "titulo": "OperaÃ§Ã£o desmantela ponto de venda no Morro da Liberdade",
                "resumo": '"Professor" foi preso com quatro pessoas durante operaÃ§Ã£o no bairro.',
                "risco": "MÃ‰DIO",
                "timestamp": (now - timedelta(hours=3)).isoformat(),
                "lido": False,
                "alvo_id": 6, "alvo_nome": "Adalberto SalomÃ£o Guedes da Silva",
                "alvo_vulgos": ["Professor", "SalomÃ£o"],
                "termo_encontrado": "professor",
                "analise_ia": None,
            },
        ]
        _salvar_alertas(_ALERTAS_PATH, alertas)

    if not os.path.exists(_ALERTAS_OSINT_PATH):
        osint = [
            {
                "id": "o1", "tipo": "sherlock", "fonte": "Sherlock â€” TikTok",
                "link": "https://www.tiktok.com/",
                "titulo": "Perfil encontrado: @carnauba_am no TikTok",
                "resumo": "Username 'carnauba' identificado em conta ativa. Bio: 'Compensa ðŸ”´âš«'. "
                          "Ãšltimo post hÃ¡ 3 dias. PossÃ­vel perfil operacional do alvo.",
                "risco": "ALTO",
                "timestamp": (now - timedelta(hours=2)).isoformat(),
                "lido": False,
                "alvo_id": 1, "alvo_nome": "Gelson de Lima CarnaÃºba", "alvo_vulgos": ["CarnaÃºba"],
                "termo_encontrado": "carnauba", "plataforma": "TikTok",
                "analise_ia": "Perfil ativo com simbologia de facÃ§Ã£o na bio. "
                              "Recomenda-se monitoramento contÃ­nuo e extraÃ§Ã£o de contatos/seguidores.",
            },
            {
                "id": "o2", "tipo": "google_dork", "fonte": "Google Dork â€” Pastebin",
                "link": "https://pastebin.com/",
                "titulo": '"MÃ£o Branca" indexado no Pastebin',
                "resumo": 'Documento indexado contÃ©m o termo "MÃ£o Branca" associado a coordenadas e '
                          "horÃ¡rios de entrega. PossÃ­vel lista operacional vazada.",
                "risco": "ALTO",
                "timestamp": (now - timedelta(hours=4)).isoformat(),
                "lido": False,
                "alvo_id": 23, "alvo_nome": "Josias da Cruz Barroso",
                "alvo_vulgos": ["MÃ£o Branca", "MB"],
                "termo_encontrado": "mÃ£o branca",
                "dork": 'site:pastebin.com "MÃ£o Branca"',
                "analise_ia": "PossÃ­vel vazamento de dados operacionais. "
                              "Prioridade mÃ¡xima â€” acionar equipe de anÃ¡lise digital.",
            },
            {
                "id": "o3", "tipo": "sherlock", "fonte": "Sherlock â€” Instagram",
                "link": "https://www.instagram.com/",
                "titulo": "Perfil encontrado: @rdk_manaus no Instagram",
                "resumo": "Username 'RDK' identificado em perfil privado. Foto de capa com referÃªncias "
                          "Ã  zona norte de Manaus. 847 seguidores.",
                "risco": "MÃ‰DIO",
                "timestamp": (now - timedelta(hours=6)).isoformat(),
                "lido": False,
                "alvo_id": 28, "alvo_nome": "Gilson Mattos Rodrigues",
                "alvo_vulgos": ["RDK", "Rei do Skunk"],
                "termo_encontrado": "rdk", "plataforma": "Instagram",
                "analise_ia": None,
            },
        ]
        _salvar_alertas(_ALERTAS_OSINT_PATH, osint)


_seed_alertas_iniciais()


# â”€â”€â”€ Seed inicial de produÃ§Ã£o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _seed_dashboard_inicial() -> None:
    if os.path.exists(_DASHBOARD_STATS_PATH):
        return
    import random as _r
    _r.seed(42)
    nucleos_docs = {
        "NI":            ["RELINT", "REPEN", "PEDIDO DE BUSCA", "MINUTA DE OFICIO", "PROJETO"],
        "NCI":           ["RELINT", "PEDIDO DE BUSCA", "RELTEC", "MINUTA DE OFICIO", "PROJETO"],
        "NBE":           ["RELINT", "RELTEC", "PROJETO"],
        "NUCADI_UPP":    ["RELATORIO INTERNO"],
        "NUCADI_COMPAJ": ["RELATORIO INTERNO"],
        "NUCADI_IPAT":   ["RELATORIO INTERNO"],
        "NUCADI_CDPM1":  ["RELATORIO INTERNO"],
        "NUCADI_CDPMII": ["RELATORIO INTERNO"],
        "NUCADI_CDF":    ["RELATORIO INTERNO"],
    }
    dados: dict = {}
    for nucleo, docs in nucleos_docs.items():
        dados[nucleo] = {}
        is_nucadi = nucleo.startswith("NUCADI")
        for m in range(12):
            dados[nucleo][str(m)] = {}
            for doc in docs:
                base = 8 if is_nucadi else 12
                dados[nucleo][str(m)][doc] = _r.randint(2, base + 1)
    os.makedirs(os.path.dirname(_DASHBOARD_STATS_PATH), exist_ok=True)
    with open(_DASHBOARD_STATS_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


_seed_dashboard_inicial()


# â”€â”€â”€ App â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

app = FastAPI(title="Agent Bastos API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5173",
        "http://127.0.0.1:5174",
    ],
    allow_methods=["POST", "GET", "OPTIONS", "PATCH", "PUT", "DELETE"],
    allow_headers=["Content-Type", "Authorization"],
)
app.include_router(liderancas_router)
app.include_router(alertas_router)
app.include_router(dashboard_router)
app.include_router(transcricao_router)

# â”€â”€â”€ Health â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}

# â”€â”€â”€ Modelos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

class PerguntaRequest(BaseModel):
    pergunta: str

    @field_validator("pergunta")
    @classmethod
    def validar_tamanho(cls, v: str) -> str:
        if len(v) > _MAX_PERGUNTA:
            raise ValueError(f"Pergunta excede {_MAX_PERGUNTA} caracteres.")
        return v


class RelatorioRequest(BaseModel):
    mes:   str
    ano:   int
    dados: dict


# â”€â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/noticias")
def noticias():
    json_path = os.path.join(PASTA_RELATORIOS, "noticias_crimes.json")
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                dados = json.load(f)
            stat = os.stat(json_path)
            noticias_list = dados.get("noticias")
            if not noticias_list and "texto" in dados:
                noticias_list = json.loads(dados["texto"]).get("noticias")
            if noticias_list:
                arquivos = [
                    {
                        "titulo":    n.get("titulo", "Sem tÃ­tulo"),
                        "resumo":    n.get("resumo", ""),
                        "link":      n.get("link", ""),
                        "imagem":    n.get("imagem", ""),
                        "data_pub":  n.get("data", ""),
                        "categoria": n.get("categoria", "CRIMES"),
                        "conteudo":  n.get("resumo", ""),
                        "arquivo":   "noticias_crimes.json",
                        "atualizado": stat.st_mtime,
                        "formato":   "estruturado",
                    }
                    for n in noticias_list
                ]
                return {"noticias": arquivos}
        except Exception:
            pass

    arquivos = []
    for caminho in glob.glob(os.path.join(PASTA_RELATORIOS, "*.txt")):
        nome = os.path.basename(caminho)
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                conteudo = f.read()
        except Exception:
            continue
        stat   = os.stat(caminho)
        titulo = (
            "Monitor Crimes AM"
            if nome == "relatorio.txt"
            else nome.replace(".txt", "").replace("_", " ").title()
        )
        arquivos.append({
            "titulo":    titulo,
            "resumo":    conteudo[:200],
            "link":      "",
            "imagem":    "",
            "data_pub":  "",
            "categoria": "CRIMES",
            "conteudo":  conteudo,
            "arquivo":   nome,
            "atualizado": stat.st_mtime,
            "formato":   "texto",
        })

    arquivos.sort(key=lambda x: x["atualizado"], reverse=True)
    return {"noticias": arquivos}


@app.post("/noticias/salvar")
async def salvar_noticias(dados: dict):
    caminho = os.path.join(PASTA_RELATORIOS, "noticias_crimes.json")
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return {"status": "salvo", "total": len(dados.get("noticias", []))}


# â”€â”€â”€ TranscriÃ§Ã£o de Ã¡udio â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


# â”€â”€â”€ Agenda de MissÃ£o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

import hashlib as _hashlib
_SENHA_CHEFE_HASH = _hashlib.sha256(b"aipen2025").hexdigest()


class AgendaLoginRequest(BaseModel):
    senha: str


class MissaoRequest(BaseModel):
    nucleo: str
    mensagem: str


@app.post("/agenda/login")
def agenda_login(req: AgendaLoginRequest):
    ok = _hashlib.sha256(req.senha.encode()).hexdigest() == _SENHA_CHEFE_HASH
    return {"ok": ok}


@app.post("/agenda/publicar")
def agenda_publicar(req: MissaoRequest):
    try:
        from modules.agenda import publicar_missao
        ok = publicar_missao(req.nucleo, req.mensagem)
        return {"ok": ok}
    except Exception as e:
        return {"ok": False, "erro": str(e)}


@app.get("/agenda/missoes")
def agenda_missoes(nucleo: str = None, limite: int = 30):
    try:
        from modules.agenda import buscar_missoes_recentes
        missoes = buscar_missoes_recentes(nucleo=nucleo, limite=limite)
        resultado = []
        for m in missoes:
            ts = m.get("timestamp")
            resultado.append({
                **m,
                "timestamp": ts.isoformat() if hasattr(ts, "isoformat") else str(ts) if ts else None,
            })
        return {"missoes": resultado}
    except Exception as e:
        return {"missoes": [], "erro": str(e)}
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ADICIONAR NO api.py â€” logo apÃ³s o endpoint GET /agenda/missoes
# Cole o bloco abaixo no arquivo api.py existente
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

class CienciaRequest(BaseModel):
    nucleo: str   # nÃºcleo que estÃ¡ acusando ciÃªncia


@app.patch("/agenda/missoes/{missao_id}/ciencia")
def agenda_acusar_ciencia(missao_id: str, req: CienciaRequest):
    """
    Agente acusa ciÃªncia de uma missÃ£o.
    Atualiza status â†’ 'ciencia' no Firestore.
    O chefe vÃª o status atualizado na prÃ³xima busca.
    """
    try:
        from modules.agenda import acusar_ciencia
        ok = acusar_ciencia(missao_id, req.nucleo)
        return {"ok": ok}
    except Exception as e:
        return {"ok": False, "erro": str(e)}

@app.get("/status")
def status():
    return {"status": "online", "version": "1.0.0", "model": "llama-3.3-70b-versatile"}


@app.get("/status/firebase")
def status_firebase():
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as _firestore
        if not firebase_admin._apps:
            cred = credentials.Certificate(_SA_KEY_PATH)
            firebase_admin.initialize_app(cred)
        db = _firestore.client()
        db.collection("missoes").limit(1).get()
        projeto = firebase_admin.get_app().project_id
        return {"ok": True, "projeto": projeto}
    except Exception as e:
        return {"ok": False, "projeto": str(e)}


# â”€â”€â”€ Firestore helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _get_firestore():
    import firebase_admin
    from firebase_admin import credentials, firestore as _firestore
    if not firebase_admin._apps:
        cred = credentials.Certificate(_SA_KEY_PATH)
        firebase_admin.initialize_app(cred)
    return _firestore.client()


def _serializar_alerta(doc) -> dict:
    d = doc.to_dict()
    d["id"] = doc.id
    ts = d.get("timestamp")
    if ts and hasattr(ts, "isoformat"):
        d["timestamp"] = ts.isoformat()
    elif ts:
        d["timestamp"] = str(ts)
    return d


# â”€â”€â”€ Alertas â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/dashboard/stats")
def get_dashboard_stats():
    if not os.path.exists(_DASHBOARD_STATS_PATH):
        _seed_dashboard_inicial()
    try:
        with open(_DASHBOARD_STATS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


@app.post("/dashboard/stats")
async def salvar_dashboard_stats(dados: dict):
    os.makedirs(os.path.dirname(_DASHBOARD_STATS_PATH), exist_ok=True)
    with open(_DASHBOARD_STATS_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    return {"status": "salvo"}


# â”€â”€â”€ ReferÃªncias â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/referencias")
def buscar_referencias(q: str = "", ano: str = "", tipo: str = ""):
    if not os.path.exists(_INDICE_PATH):
        return {"documentos": [], "anos": [], "total": 0, "total_indexados": 0}
    try:
        with open(_INDICE_PATH, "r", encoding="utf-8") as f:
            indice = json.load(f)
    except Exception:
        return {"documentos": [], "anos": [], "total": 0, "total_indexados": 0}

    docs  = indice.get("documentos", [])
    anos  = sorted({d.get("ano", "") for d in docs if d.get("ano")}, reverse=True)
    termo = q.strip().upper()

    resultado = []
    for doc in docs:
        if ano  and doc.get("ano")  != ano:  continue
        if tipo and doc.get("tipo") != tipo:  continue
        if termo:
            assunto = (doc.get("assunto") or "").upper()
            numero  = doc.get("numero", "")
            if termo not in assunto and termo not in numero:
                continue
        resultado.append(doc)

    resultado.sort(key=lambda d: (d.get("ano", ""), d.get("numero", "")), reverse=True)
    return {
        "documentos":      resultado,
        "anos":            anos,
        "total":           len(resultado),
        "total_indexados": len(docs),
    }


# â”€â”€â”€ Lista Negra (Google Drive) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

from googleapiclient.discovery import build as _gdrive_build
from googleapiclient.http import MediaIoBaseDownload as _MediaIoBaseDownload
from google.oauth2 import service_account as _sa
import openpyxl as _openpyxl

# ── Drive service (Passo 1 da refatoração) ────────────────────────────────────
# As funções _gdrive_service, _upload_json_drive e _baixar_json_drive foram
# movidas para services/drive_service.py. Os aliases abaixo mantêm compatibilidade
# com todo o código existente sem alterar nenhuma chamada.
from services.drive_service import (
    get_service        as _gdrive_service_new,
    upload_json        as _upload_json_drive,
    download_json      as _baixar_json_drive,
    download_bytes     as _download_bytes_drive,
)


def _mascarar_cpf(cpf: str) -> str:
    """753.164.392-87 â†’ 753.***.***-**"""
    if not cpf:
        return ""
    limpo = re.sub(r"\D", "", str(cpf))
    if len(limpo) < 3:
        return "***.***.***-**"
    return f"{limpo[:3]}.***.***-**"


def _baixar_xlsx_drive() -> bytes:
    """Alias de compatibilidade → services.drive_service.download_bytes()"""
    return _download_bytes_drive(_LISTA_NEGRA_FILE_ID)


def _ler_lista_negra() -> list:
    """LÃª todas as abas do .xlsx e retorna lista unificada ordenada A-Z por nome."""
    xlsx_bytes = _baixar_xlsx_drive()
    wb         = _openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    registros  = []

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Detecta linha do cabeÃ§alho (procura "NOME" nas primeiras 5 linhas)
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            vals = [str(v).upper() for v in row if v]
            if "NOME" in vals:
                header_idx = i
                break

        headers = [str(v).strip().upper() if v else "" for v in rows[header_idx]]

        col_map = {}
        for i, h in enumerate(headers):
            if "NOME" in h:                               col_map["nome"]       = i
            elif h in ("N", "NÂ°", "NÂº") or "NUM" in h:   col_map["numero"]     = i
            elif "REFER" in h:                            col_map["referencia"] = i
            elif "SITUA" in h:                            col_map["situacao"]   = i
            elif "UNID" in h:                             col_map["unidade"]    = i
            elif "EMPRES" in h:                           col_map["empresa"]    = i
            elif "DATA" in h:                             col_map["data"]       = i
            elif "CPF" in h:                              col_map["cpf"]        = i
            elif "DESC" in h or "OBS" in h:               col_map["descricao"]  = i

        if "nome" not in col_map:
            continue

        for row in rows[header_idx + 1:]:
            nome = row[col_map["nome"]] if col_map.get("nome") is not None else None
            if not nome or str(nome).strip() == "":
                continue

            def _get(key, _row=row, _col=col_map):
                idx = _col.get(key)
                if idx is None or idx >= len(_row):
                    return ""
                val = _row[idx]
                if val is None:
                    return ""
                if hasattr(val, "strftime"):
                    return val.strftime("%d/%m/%Y")
                return str(val).strip()

            registros.append({
                "numero":     _get("numero"),
                "nome":       str(nome).strip().upper(),
                "referencia": _get("referencia"),
                "situacao":   _get("situacao").upper(),
                "unidade":    _get("unidade"),
                "empresa":    _get("empresa"),
                "data":       _get("data"),
                "cpf":        _mascarar_cpf(_get("cpf")),
                "descricao":  _get("descricao"),
                "ano":        sheet_name,
            })

    registros.sort(key=lambda x: x["nome"])
    return registros


@app.get("/lista-negra")
def lista_negra():
    """
    Retorna lista de servidores da planilha Caveirinha (Google Drive).
    CPF mascarado â€” apenas os 3 primeiros dÃ­gitos visÃ­veis. LGPD compliant.
    """
    try:
        registros = _ler_lista_negra()
        return {"total": len(registros), "registros": registros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao acessar planilha: {e}")


# â”€â”€â”€ Entry point â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€



# --- Download DOCX por file_id (Google Drive) ---
@app.get("/referencias/download/docx/{file_id}")
def download_referencia_docx(file_id: str):
    import io as _io
    from googleapiclient.discovery import build as _gdrive_build
    from googleapiclient.http import MediaIoBaseDownload as _MediaDL
    from google.oauth2 import service_account as _sa2
    from fastapi.responses import StreamingResponse
    try:
        creds = _sa2.Credentials.from_service_account_file(
            _SA_KEY_PATH, scopes=["https://www.googleapis.com/auth/drive.readonly"])
        service = _gdrive_build("drive", "v3", credentials=creds)
        request = service.files().get_media(fileId=file_id)
        buf = _io.BytesIO()
        dl = _MediaDL(buf, request)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf.seek(0)
        # Busca nome real no indice
        import json as _json
        nome = file_id
        try:
            with open(_INDICE_PATH, encoding='utf-8') as _f:
                _idx = _json.load(_f)
            _doc = next((d for d in _idx['documentos'] if d.get('file_id') == file_id), None)
            if _doc:
                tipo = _doc.get('tipo', '')
                num = _doc.get('numero', '')
                ano = _doc.get('ano', '')
                fmt = _doc.get('formato', 'docx')
                nome = f"{tipo}_{num}_{ano}.{fmt}"
        except Exception:
            pass
        mt = "application/vnd.openxmlformats-officedocument.wordprocessingml.document" if nome.endswith('docx') else "application/pdf"
        return StreamingResponse(buf, media_type=mt,
            headers={"Content-Disposition": f"attachment; filename={nome}"})
    except Exception as e:
        return {"erro": str(e)}

# --- Download PDF por file_id (Google Drive - converte via API) ---
@app.get("/referencias/download/pdf/{file_id}")
def download_referencia_pdf(file_id: str):
    import io as _io
    from googleapiclient.discovery import build as _gdrive_build
    from googleapiclient.http import MediaIoBaseDownload as _MediaDL
    from google.oauth2 import service_account as _sa2
    from fastapi.responses import StreamingResponse
    try:
        creds = _sa2.Credentials.from_service_account_file(
            _SA_KEY_PATH, scopes=["https://www.googleapis.com/auth/drive"])
        service = _gdrive_build("drive", "v3", credentials=creds)
        # Importa o arquivo como Google Doc
        meta = {"name": file_id, "mimeType": "application/vnd.google-apps.document"}
        media_body = service.files().get_media(fileId=file_id)
        buf_orig = _io.BytesIO()
        dl = _MediaDL(buf_orig, media_body)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf_orig.seek(0)
        from googleapiclient.http import MediaIoBaseUpload
        uploaded = service.files().create(
            body=meta,
            media_body=MediaIoBaseUpload(buf_orig,
                mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            fields="id"
        ).execute()
        temp_id = uploaded["id"]
        # Exporta como PDF
        pdf_request = service.files().export_media(fileId=temp_id, mimeType="application/pdf")
        pdf_buf = _io.BytesIO()
        dl2 = _MediaDL(pdf_buf, pdf_request)
        done2 = False
        while not done2:
            _, done2 = dl2.next_chunk()
        # Remove arquivo temporario
        service.files().delete(fileId=temp_id).execute()
        pdf_buf.seek(0)
        return StreamingResponse(pdf_buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={file_id}.pdf"})
    except Exception as e:
        return {"erro": str(e)}
_IMG_MIME_MAP = {
    "image/jpeg": ".jpg",
    "image/png":  ".png",
    "image/webp": ".webp",
    "image/gif":  ".gif",
}

@app.post("/decifrar")
async def decifrar_missiva(
    imagem: UploadFile = File(...),
    tipo_documento: str = Form("desconhecido"),
    contexto_extra: str = Form(""),
):
    if imagem.content_type not in _IMG_MIME_MAP:
        raise HTTPException(status_code=415, detail=f"Formato nÃ£o suportado: {imagem.content_type}")
    dados = await imagem.read()
    if len(dados) > _MAX_AUDIO_BYTES:
        raise HTTPException(status_code=413, detail="Imagem excede o limite de 25MB.")
    try:
        tipo_enum = TipoDocumento(tipo_documento.lower())
    except ValueError:
        tipo_enum = TipoDocumento.DESCONHECIDO
    return transcrever_documento_bytes(
        dados=dados,
        media_type=imagem.content_type,
        nome_arquivo=imagem.filename or "documento",
        tipo_documento=tipo_enum,
        contexto_extra=contexto_extra,
    )
# rotas dashboard migradas para routers/dashboard_router.py (Passo 4)

# --- Lideranças por Unidade ---
@app.get("/ocupacao")
def get_ocupacao():
    try:
        return _baixar_ocupacao_drive()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Inteligência de Grupos — Snapshots e KPIs ---
_HISTORICO_FOLDER_ID = "1-hQE2t9P7Kpk31V5oKCUoylM03nCNhT_"

# ── _gdrive_service, _upload_json_drive, _baixar_json_drive ──────────────────
# Movidas para services/drive_service.py (Passo 1 da refatoração).
# Alias local mantém compatibilidade com todo o código abaixo sem alterar chamadas.
def _gdrive_service():
    """Alias de compatibilidade → services.drive_service.get_service()"""
    return _gdrive_service_new(readonly=False)

def _salvar_snapshot_automatico():
    mes_atual = datetime.now().strftime("%Y-%m")
    nome = f"snapshot_{mes_atual}.json"
    existente = _baixar_json_drive(nome, _HISTORICO_FOLDER_ID)
    if existente:
        return False
    try:
        ocupacao = _baixar_ocupacao_drive()
        snapshot = {"mes": mes_atual, "gerado_em": datetime.now().isoformat(), "dados": ocupacao}
        _upload_json_drive(nome, snapshot, _HISTORICO_FOLDER_ID)
        indice = _baixar_json_drive("indice.json", _HISTORICO_FOLDER_ID) or {"meses": []}
        if mes_atual not in indice["meses"]:
            indice["meses"].append(mes_atual)
            indice["meses"].sort()
        _upload_json_drive("indice.json", indice, _HISTORICO_FOLDER_ID)
        return True
    except Exception:
        return False

@app.get("/historico/indice")
def get_indice():
    try:
        indice = _baixar_json_drive("indice.json", _HISTORICO_FOLDER_ID)
        return indice or {"meses": []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/historico/{mes}")
def get_historico_mes(mes: str):
    try:
        snap = _baixar_json_drive(f"snapshot_{mes}.json", _HISTORICO_FOLDER_ID)
        if not snap:
            raise HTTPException(status_code=404, detail=f"Snapshot {mes} nao encontrado")
        return snap
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/kpis")
def get_kpis():
    try:
        indice = _baixar_json_drive("indice.json", _HISTORICO_FOLDER_ID) or {"meses": []}
        meses = sorted(indice.get("meses", []))
        historico = []
        for mes in meses[-6:]:
            snap = _baixar_json_drive(f"snapshot_{mes}.json", _HISTORICO_FOLDER_ID)
            if snap:
                historico.append({"mes": mes, "dados": snap.get("dados", {})})
        series = {}
        for h in historico:
            mes = h["mes"]
            contagem = {}
            for u_dados in h["dados"].get("unidades", {}).values():
                for p in (u_dados.get("pavilhoes") or u_dados.get("pavs") or {}).values():
                    g = p.get("grupo") or p.get("g", "")
                    contagem[g] = contagem.get(g, 0) + 1
            series[mes] = contagem
        alertas = []
        meses_list = sorted(series.keys())
        if len(meses_list) >= 2:
            atual = series[meses_list[-1]]
            anterior = series[meses_list[-2]]
            for grupo, qtd in atual.items():
                qtd_ant = anterior.get(grupo, 0)
                if qtd_ant > 0:
                    variacao = ((qtd - qtd_ant) / qtd_ant) * 100
                    if abs(variacao) >= 20:
                        alertas.append({"grupo": grupo, "variacao": round(variacao, 1), "atual": qtd, "anterior": qtd_ant})
        return {"series": series, "alertas": alertas, "meses": meses_list}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# startup_snapshot removido — snapshot agora é manual via /snapshot/forcar

@app.post("/snapshot/forcar")
def forcar_snapshot():
    try:
        mes_atual = datetime.now().strftime("%Y-%m")
        ocupacao = _baixar_ocupacao_drive()
        snapshot = {"mes": mes_atual, "gerado_em": datetime.now().isoformat(), "dados": ocupacao}
        _upload_json_drive(f"snapshot_{mes_atual}.json", snapshot, _HISTORICO_FOLDER_ID)
        indice = _baixar_json_drive("indice.json", _HISTORICO_FOLDER_ID) or {"meses": []}
        if mes_atual not in indice["meses"]:
            indice["meses"].append(mes_atual)
            indice["meses"].sort()
        _upload_json_drive("indice.json", indice, _HISTORICO_FOLDER_ID)
        return {"ok": True, "mes": mes_atual}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/drive")
def debug_drive():
    import traceback
    try:
        service = _gdrive_service()
        results = service.files().list(
            q=f"'{_HISTORICO_FOLDER_ID}' in parents and trashed=false",
            fields="files(id, name)"
        ).execute()
        return {"ok": True, "arquivos": results.get("files", [])}
    except Exception as e:
        return {"ok": False, "erro": str(e), "trace": traceback.format_exc()}

@app.get("/debug/snapshot-erro")
def debug_snapshot_erro():
    import traceback
    try:
        mes_atual = datetime.now().strftime("%Y-%m")
        ocupacao = _baixar_ocupacao_drive()
        snapshot = {"mes": mes_atual, "gerado_em": datetime.now().isoformat(), "dados": ocupacao}
        _upload_json_drive(f"snapshot_{mes_atual}.json", snapshot, _HISTORICO_FOLDER_ID)
        return {"ok": True, "mes": mes_atual}
    except Exception as e:
        return {"ok": False, "erro": str(e), "trace": traceback.format_exc()}

# --- Snapshots: Drive (fonte de verdade) + Local (cache/fallback) ---
import os as _os
_SNAPSHOTS_DIR = _os.path.join(BASE_DIR, "data", "snapshots")
_os.makedirs(_SNAPSHOTS_DIR, exist_ok=True)


def _computar_series_kpis(meses_snaps: list) -> dict:
    """Recebe lista de {'mes': str, 'dados': dict} e retorna series + alertas."""
    series = {}
    for item in meses_snaps:
        mes = item["mes"]
        contagem = {}
        for u_dados in item.get("dados", {}).get("unidades", {}).values():
            for p in (u_dados.get("pavilhoes") or u_dados.get("pavs") or {}).values():
                g = p.get("grupo") or p.get("g", "")
                if g:
                    contagem[g] = contagem.get(g, 0) + 1
        series[mes] = contagem

    alertas = []
    meses_list = sorted(series.keys())
    if len(meses_list) >= 2:
        atual = series[meses_list[-1]]
        anterior = series[meses_list[-2]]
        for grupo, qtd in atual.items():
            qtd_ant = anterior.get(grupo, 0)
            if qtd_ant > 0:
                variacao = ((qtd - qtd_ant) / qtd_ant) * 100
                if abs(variacao) >= 20:
                    alertas.append({
                        "grupo": grupo,
                        "variacao": round(variacao, 1),
                        "atual": qtd,
                        "anterior": qtd_ant,
                    })
    return {"series": series, "alertas": alertas, "meses": meses_list}


def _salvar_snapshot_completo():
    """
    Salva snapshot do mês atual TANTO no Drive quanto localmente.
    Drive = fonte de verdade e histórico persistente.
    Local = cache offline/fallback.
    Idempotente: não sobrescreve se já existir no Drive no mesmo mês.
    """
    mes_atual = datetime.now().strftime("%Y-%m")
    nome_arquivo = f"snapshot_{mes_atual}.json"

    existente_drive = _baixar_json_drive(nome_arquivo, _HISTORICO_FOLDER_ID)
    if existente_drive:
        return {"criado": False, "destino": "drive", "mes": mes_atual}

    try:
        ocupacao = _baixar_ocupacao_drive()
        snapshot = {
            "mes": mes_atual,
            "gerado_em": datetime.now().isoformat(),
            "dados": ocupacao,
        }

        # Drive
        _upload_json_drive(nome_arquivo, snapshot, _HISTORICO_FOLDER_ID)
        indice = _baixar_json_drive("indice.json", _HISTORICO_FOLDER_ID) or {"meses": []}
        if mes_atual not in indice["meses"]:
            indice["meses"].append(mes_atual)
            indice["meses"].sort()
        _upload_json_drive("indice.json", indice, _HISTORICO_FOLDER_ID)

        # Local (cache)
        path_local = _os.path.join(_SNAPSHOTS_DIR, nome_arquivo)
        with open(path_local, "w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
        indice_local_path = _os.path.join(_SNAPSHOTS_DIR, "indice.json")
        ind_local = json.loads(open(indice_local_path).read()) if _os.path.exists(indice_local_path) else {"meses": []}
        if mes_atual not in ind_local["meses"]:
            ind_local["meses"].append(mes_atual)
            ind_local["meses"].sort()
        with open(indice_local_path, "w", encoding="utf-8") as f:
            json.dump(ind_local, f, ensure_ascii=False, indent=2)

        return {"criado": True, "destino": "drive+local", "mes": mes_atual}
    except Exception as e:
        return {"criado": False, "erro": str(e)}


# startup_snapshot_unificado removido — snapshot agora é manual via /snapshot/forcar


# ═══════════════════════════════════════════════════════════════════════════════
# LIDERANÇAS DE PAVILHÕES
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/liderancas-estrutura")
def get_estrutura():
    """Estrutura física completa com celas + facções + cargos por facção."""
    return {
        "estrutura":         estrutura_com_celas(),
        "faccoes":           FACCOES,
        "cargos_por_faccao": CARGOS_POR_FACCAO,
    }


@app.get("/liderancas/{unidade}")
def get_liderancas_unidade(unidade: str):
    """Líderes de uma unidade agrupados por pavilhão → ala → cela."""
    if unidade not in ESTRUTURA:
        raise HTTPException(status_code=404, detail=f"Unidade '{unidade}' não encontrada.")
    return {
        "unidade":  unidade,
        "label":    ESTRUTURA[unidade]["label"],
        "pavilhoes": listar_por_unidade(unidade),
    }


@app.post("/liderancas")
async def post_lider(
    unidade:    str = Form(...),
    pavilhao:   str = Form(...),
    ala:        str = Form(...),
    cela:       str = Form(""),
    faccao:     str = Form(...),
    cargo:      str = Form(...),
    nome:       str = Form(""),
    vulgo:      str = Form(""),
    observacao: str = Form(""),
    foto: UploadFile = File(None),
):
    """Cria um novo líder. Foto é opcional."""
    dados = {
        "unidade": unidade, "pavilhao": pavilhao, "ala": ala, "cela": cela,
        "faccao": faccao,   "cargo": cargo,
        "nome": nome or None, "vulgo": vulgo or None,
        "observacao": observacao or None,
        "foto_ext": None,
    }
    lider = criar_lider(dados)

    if foto and foto.filename:
        ext      = os.path.splitext(foto.filename)[1] or ".jpg"
        conteudo = await foto.read()
        if len(conteudo) > 5 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Foto maior que 5 MB.")
        ext_salva = salvar_foto(lider["id"], conteudo, ext)
        lider     = atualizar_lider(lider["id"], {"foto_ext": ext_salva})

    return lider


@app.put("/liderancas/{lider_id}")
async def put_lider(
    lider_id:   str,
    unidade:    str = Form(...),
    pavilhao:   str = Form(...),
    ala:        str = Form(...),
    cela:       str = Form(""),
    faccao:     str = Form(...),
    cargo:      str = Form(...),
    nome:       str = Form(""),
    vulgo:      str = Form(""),
    observacao: str = Form(""),
    foto: UploadFile = File(None),
):
    """Atualiza líder. Nova foto substitui a anterior."""
    if not buscar_lider(lider_id):
        raise HTTPException(status_code=404, detail="Líder não encontrado.")

    dados = {
        "unidade": unidade, "pavilhao": pavilhao, "ala": ala, "cela": cela,
        "faccao": faccao,   "cargo": cargo,
        "nome": nome or None, "vulgo": vulgo or None,
        "observacao": observacao or None,
    }

    if foto and foto.filename:
        ext      = os.path.splitext(foto.filename)[1] or ".jpg"
        conteudo = await foto.read()
        if len(conteudo) > 5 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Foto maior que 5 MB.")
        ext_salva      = salvar_foto(lider_id, conteudo, ext)
        dados["foto_ext"] = ext_salva

    return atualizar_lider(lider_id, dados)


@app.delete("/liderancas/{lider_id}")
def delete_lider(lider_id: str):
    """Remove líder e sua foto do disco."""
    if not deletar_lider(lider_id):
        raise HTTPException(status_code=404, detail="Líder não encontrado.")
    return {"ok": True}


@app.get("/liderancas-foto/{lider_id}")
def get_foto_lider(lider_id: str):
    """Serve a foto binária do líder com mime type correto."""
    lider = buscar_lider(lider_id)
    if not lider or not lider.get("foto_ext"):
        raise HTTPException(status_code=404, detail="Foto não encontrada.")
    conteudo = carregar_foto(lider_id, lider["foto_ext"])
    if not conteudo:
        raise HTTPException(status_code=404, detail="Arquivo de foto ausente no disco.")
    ext  = lider["foto_ext"].lstrip(".")
    mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png",  "webp": "image/webp"}.get(ext, "image/jpeg")
    return Response(content=conteudo, media_type=mime)

if __name__ == "__main__":
    uvicorn.run("api:app", host="127.0.0.1", port=8000, reload=False)
