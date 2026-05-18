"""
routers/referencias_router.py
─────────────────────────────────────────────────────────────────────────────
Rotas HTTP de referências documentais, lista negra e downloads do Drive.

Rotas registradas:
  GET /referencias                        → busca no índice local JSON
  GET /lista-negra                        → planilha xlsx do Google Drive (LGPD)
  GET /referencias/download/docx/{id}    → download direto do arquivo
  GET /referencias/download/pdf/{id}     → converte docx→pdf via Drive API

LGPD:
  CPF mascarado — apenas 3 primeiros dígitos visíveis (_mascarar_cpf).
"""

import io
import json
import os
import re

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from services.drive_service import download_bytes, get_service

router = APIRouter(tags=["referencias"])

# ─── Constantes ──────────────────────────────────────────────────────────────

BASE_DIR             = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SA_KEY_PATH         = os.path.join(BASE_DIR, "serviceAccountKey.json")
_INDICE_PATH         = os.path.join(BASE_DIR, "indice_documentos.json")
_LISTA_NEGRA_FILE_ID = "1G6eFhb0jnD38iWU_SLDIFkGtOB0ngHjh"


# ─── Helpers privados ─────────────────────────────────────────────────────────

def _mascarar_cpf(cpf: str) -> str:
    """753.164.392-87 → 753.***.***-** (LGPD: exibe só 3 primeiros dígitos)"""
    if not cpf:
        return ""
    limpo = re.sub(r"\D", "", str(cpf))
    if len(limpo) < 3:
        return "***.***.***-**"
    return f"{limpo[:3]}.***.***-**"


def _ler_lista_negra() -> list:
    """
    Baixa o xlsx da Lista Negra do Drive e retorna registros ordenados A-Z.
    Detecta automaticamente a linha de cabeçalho (procura 'NOME' nas 5 primeiras).
    CPF mascarado antes de retornar — nunca expõe dado completo.
    """
    import openpyxl
    xlsx_bytes = download_bytes(_LISTA_NEGRA_FILE_ID)
    wb         = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    registros  = []

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header_idx = 0
        for i, row in enumerate(rows[:5]):
            vals = [str(v).upper() for v in row if v]
            if "NOME" in vals:
                header_idx = i
                break

        headers = [str(v).strip().upper() if v else "" for v in rows[header_idx]]
        col_map = {}
        for i, h in enumerate(headers):
            if "NOME" in h:                             col_map["nome"]       = i
            elif h in ("N", "N°", "Nº") or "NUM" in h: col_map["numero"]     = i
            elif "REFER" in h:                          col_map["referencia"] = i
            elif "SITUA" in h:                          col_map["situacao"]   = i
            elif "UNID" in h:                           col_map["unidade"]    = i
            elif "EMPRES" in h:                         col_map["empresa"]    = i
            elif "DATA" in h:                           col_map["data"]       = i
            elif "CPF" in h:                            col_map["cpf"]        = i
            elif "DESC" in h or "OBS" in h:             col_map["descricao"]  = i

        if "nome" not in col_map:
            continue

        for row in rows[header_idx + 1:]:
            nome = row[col_map["nome"]] if col_map.get("nome") is not None else None
            if not nome or str(nome).strip() == "":
                continue

            def _get(key, _row=row, _col=col_map):
                idx = _col.get(key)
                if idx is None or idx >= len(_row):
                    return ""
                val = _row[idx]
                if val is None:
                    return ""
                if hasattr(val, "strftime"):
                    return val.strftime("%d/%m/%Y")
                return str(val).strip()

            registros.append({
                "numero":     _get("numero"),
                "nome":       str(nome).strip().upper(),
                "referencia": _get("referencia"),
                "situacao":   _get("situacao").upper(),
                "unidade":    _get("unidade"),
                "empresa":    _get("empresa"),
                "data":       _get("data"),
                "cpf":        _mascarar_cpf(_get("cpf")),
                "descricao":  _get("descricao"),
                "ano":        sheet_name,
            })

    registros.sort(key=lambda x: x["nome"])
    return registros


# ─── Rotas ───────────────────────────────────────────────────────────────────

@router.get("/referencias")
def buscar_referencias(q: str = "", ano: str = "", tipo: str = ""):
    if not os.path.exists(_INDICE_PATH):
        return {"documentos": [], "anos": [], "total": 0, "total_indexados": 0}
    try:
        with open(_INDICE_PATH, "r", encoding="utf-8") as f:
            indice = json.load(f)
    except Exception:
        return {"documentos": [], "anos": [], "total": 0, "total_indexados": 0}

    docs  = indice.get("documentos", [])
    anos  = sorted({d.get("ano", "") for d in docs if d.get("ano")}, reverse=True)
    termo = q.strip().upper()

    resultado = []
    for doc in docs:
        if ano  and doc.get("ano")  != ano:  continue
        if tipo and doc.get("tipo") != tipo:  continue
        if termo:
            assunto = (doc.get("assunto") or "").upper()
            numero  = doc.get("numero", "")
            if termo not in assunto and termo not in numero:
                continue
        resultado.append(doc)

    resultado.sort(key=lambda d: (d.get("ano", ""), d.get("numero", "")), reverse=True)
    return {
        "documentos":      resultado,
        "anos":            anos,
        "total":           len(resultado),
        "total_indexados": len(docs),
    }


@router.get("/lista-negra")
def lista_negra():
    """CPF mascarado — LGPD compliant."""
    try:
        registros = _ler_lista_negra()
        return {"total": len(registros), "registros": registros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao acessar planilha: {e}")


@router.get("/referencias/download/docx/{file_id}")
def download_referencia_docx(file_id: str):
    try:
        buf = io.BytesIO(download_bytes(file_id))
        nome = file_id
        try:
            with open(_INDICE_PATH, encoding="utf-8") as f:
                idx = json.load(f)
            doc = next((d for d in idx["documentos"] if d.get("file_id") == file_id), None)
            if doc:
                nome = f"{doc.get('tipo','')}_{doc.get('numero','')}_{doc.get('ano','')}.{doc.get('formato','docx')}"
        except Exception:
            pass
        mt = (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            if nome.endswith("docx") else "application/pdf"
        )
        return StreamingResponse(buf, media_type=mt,
            headers={"Content-Disposition": f"attachment; filename={nome}"})
    except Exception as e:
        return {"erro": str(e)}


@router.get("/referencias/download/pdf/{file_id}")
def download_referencia_pdf(file_id: str):
    """Converte DOCX → PDF via Google Drive API (upload temporário + export)."""
    try:
        from googleapiclient.http import MediaIoBaseUpload
        service  = get_service(readonly=False)
        buf_orig = io.BytesIO(download_bytes(file_id))
        uploaded = service.files().create(
            body={"name": file_id, "mimeType": "application/vnd.google-apps.document"},
            media_body=MediaIoBaseUpload(buf_orig,
                mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            fields="id",
        ).execute()
        temp_id = uploaded["id"]
        try:
            from googleapiclient.http import MediaIoBaseDownload
            pdf_req = service.files().export_media(fileId=temp_id, mimeType="application/pdf")
            pdf_buf = io.BytesIO()
            dl = MediaIoBaseDownload(pdf_buf, pdf_req)
            done = False
            while not done:
                _, done = dl.next_chunk()
            pdf_buf.seek(0)
        finally:
            service.files().delete(fileId=temp_id).execute()
        return StreamingResponse(pdf_buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={file_id}.pdf"})
    except Exception as e:
        return {"erro": str(e)}
